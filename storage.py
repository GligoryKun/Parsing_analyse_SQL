import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import json
import csv
import sqlite3

def save_SQLite(data_list, path):
    connection = sqlite3.connect(path / 'vacancy_SQL.db')
    cursor = connection.cursor()
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS vacancy (
    id INTEGER PRIMARY KEY,
    title TEXT,
    salary TEXT,
    experience TEXT,
    company TEXT,
    link TEXT
    )
    ''')
    cursor.execute('DELETE FROM vacancy')
    for data in data_list:
        cursor.execute('INSERT INTO vacancy (title, salary, experience, company, link) VALUES (?,?,?,?,?)', data)
    connection.commit()

    connection.close()

def save_CSV(file_name, header, data):
    with open(file_name, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(header)
        writer.writerows(data)

def save_josn():
    with open('vacancy.josn', 'w', encoding='utf-8') as fl:
        pass

def save_excel(filename, dict1, dict2):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # desing settings
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Headers dict
    tables = [
        {"data": dict1, "title": "Salary Metrics", "col": 1},
        {"data": dict2, "title": "Experience Stats", "col": 4}
    ]

    for table in tables:
        data = table["data"]
        start_col = table["col"]

        # 1. Header
        ws.cell(row=1, column=start_col, value=table["title"]).font = Font(bold=True, size=12)

        # 2. dict data selecting
        current_row = 2
        for key, value in data.items():
            # left cell (key - vertical cell)
            key_cell = ws.cell(row=current_row, column=start_col, value=key)
            key_cell.font = header_font
            key_cell.fill = header_fill
            key_cell.border = thin_border

            # ringh cell (value)
            val_cell = ws.cell(row=current_row, column=start_col + 1, value=value)
            val_cell.border = thin_border
            val_cell.alignment = Alignment(horizontal='center')

            current_row += 1

        # auto-set width of column for a keys
        ws.column_dimensions[openpyxl.utils.get_column_letter(start_col)].width = 20

    wb.save(filename)
